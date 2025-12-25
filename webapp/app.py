# -*- coding: utf-8 -*-
"""
SwissCapital - Веб-платформа генерации справок о ссудной задолженности
"""

import os
import sys
import uuid
import shutil
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List

from fastapi import FastAPI, Request, UploadFile, File, Form, Depends, HTTPException, status
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.security import HTTPBasic, HTTPBasicCredentials
import secrets

# Добавляем путь к scripts для импорта модулей
sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from num2text import number_to_text, format_number_with_text

# ============================================================================
# КОНФИГУРАЦИЯ
# ============================================================================

APP_DIR = Path(__file__).parent
PROJECT_DIR = APP_DIR.parent  # Корневая папка проекта
UPLOAD_DIR = APP_DIR / "uploads"
GENERATED_DIR = APP_DIR / "generated"
TEMPLATES_DIR = APP_DIR / "templates"
STATIC_DIR = APP_DIR / "static"

# Создаём директории если их нет
UPLOAD_DIR.mkdir(exist_ok=True)
GENERATED_DIR.mkdir(exist_ok=True)

# Регистрация шрифтов для PDF с поддержкой кириллицы
import platform

PDF_FONT = 'Helvetica'
PDF_FONT_BOLD = 'Helvetica-Bold'

# Определяем ОС и используем соответствующие шрифты
if platform.system() == 'Windows':
    try:
        pdfmetrics.registerFont(TTFont('Arial', 'C:/Windows/Fonts/arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
        PDF_FONT = 'Arial'
        PDF_FONT_BOLD = 'Arial-Bold'
        print("✓ Registered Arial fonts for Windows")
    except Exception as e:
        print(f"⚠ Arial registration failed: {e}")
else:
    # Linux/Docker
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
        PDF_FONT = 'DejaVu'
        PDF_FONT_BOLD = 'DejaVu-Bold'
        print("✓ Registered DejaVu fonts for Linux")
    except Exception as e:
        print(f"⚠ DejaVu registration failed: {e}, using Helvetica (no Cyrillic)")

# Авторизация
USERS = {
    "Kirito": "Kirito"
}

# Данные компании
COMPANY = {
    'name': 'SwissCapital',
    'address': 'Республика Казахстан, г. Алматы, пр. Достык, 188',
    'phone': '+7 700 836 78 13'
}

# Путь к логотипу
LOGO_PATH = PROJECT_DIR / "assetslogo.png"

# ============================================================================
# ПРИЛОЖЕНИЕ
# ============================================================================

app = FastAPI(title="SwissCapital Certificate Generator")
security = HTTPBasic()

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
templates = Jinja2Templates(directory=TEMPLATES_DIR)

# История генераций (в памяти, для production используйте БД)
generation_history = []


# ============================================================================
# АВТОРИЗАЦИЯ
# ============================================================================

def verify_credentials(credentials: HTTPBasicCredentials = Depends(security)):
    """Проверка логина и пароля"""
    username = credentials.username
    password = credentials.password

    if username in USERS and secrets.compare_digest(password, USERS[username]):
        return username

    raise HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Неверный логин или пароль",
        headers={"WWW-Authenticate": "Basic"},
    )


# ============================================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================================

def format_date_russian(date_str):
    """Форматирование даты в русском формате"""
    months = {
        1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля',
        5: 'мая', 6: 'июня', 7: 'июля', 8: 'августа',
        9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
    }
    try:
        if isinstance(date_str, datetime):
            dt = date_str
        else:
            for fmt in ['%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']:
                try:
                    dt = datetime.strptime(str(date_str), fmt)
                    break
                except ValueError:
                    continue
            else:
                return str(date_str)
        return f"{dt.day} {months[dt.month]} {dt.year} г."
    except:
        return str(date_str)


def get_column_mapping_info(file_path: Path) -> dict:
    """Получить информацию о маппинге столбцов Excel"""
    wb = load_workbook(file_path)
    ws = wb.active

    # Читаем заголовки
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip()] = col

    wb.close()

    return {
        "found_columns": list(headers.keys()),
        "total_columns": len(headers)
    }


def validate_iin(iin: str) -> dict:
    """Валидация ИИН клиента"""
    if not iin:
        return {'valid': False, 'error': 'ИИН отсутствует'}

    # Удаляем пробелы
    iin_clean = iin.strip()

    # Проверка, что содержит только цифры
    if not iin_clean.isdigit():
        return {'valid': False, 'error': f'ИИН содержит недопустимые символы: {iin_clean}'}

    # Проверка длины
    if len(iin_clean) != 12:
        return {'valid': False, 'error': f'ИИН должен содержать 12 цифр, найдено: {len(iin_clean)}'}

    return {'valid': True, 'error': None}


def read_excel_data(file_path: Path) -> List[dict]:
    """Чтение данных из Excel файла"""
    wb = load_workbook(file_path)
    ws = wb.active

    # Маппинг заголовков
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip().lower()] = col

    # Логируем найденные заголовки (в логи сервера)
    # print отключены из-за проблем с кодировкой Windows

    # Точный маппинг заголовков (в нижнем регистре) на поля
    column_map = {
        # Номер договора
        'номер договора': 'contract_number',
        'номера договора': 'contract_number',
        '№ договора': 'contract_number',
        '№ номера договора': 'contract_number',
        'договор №': 'contract_number',
        'договор': 'contract_number',
        # Игнорируем столбец с порядковым номером
        '№': 'ignore',
        'п/п': 'ignore',
        '№ п/п': 'ignore',
        # Дата договора
        'дата договора': 'contract_date',
        'даты договора': 'contract_date',
        'дата': 'contract_date',
        # ФИО
        'фио': 'client_name',
        'фио клиента': 'client_name',
        'клиент': 'client_name',
        'заемщик': 'client_name',
        # ИИН
        'иин': 'iin',
        'иин клиента': 'iin',
        'iin': 'iin',
        # Основной долг
        'основной долг': 'principal',
        'сумма од': 'principal',
        'сумма основного долга': 'principal',
        'од': 'principal',
        # Вознаграждение (проценты)
        'вознаграждение': 'reward',
        'сумма вознаграждения': 'reward',
        'сумма процентов': 'reward',
        'проценты': 'reward',
        # Отсроченные проценты / поступления
        'отсроченные проценты': 'deferred_interest',
        'сумма отсроченных процентов': 'deferred_interest',
        'отсроч проценты': 'deferred_interest',
        'отсроч. проценты': 'deferred_interest',
        'отсроченн проценты': 'deferred_interest',
        'отсроченн. проценты': 'deferred_interest',
        'отсроченные поступления': 'deferred_interest',
        'сумма отсроченных поступлений': 'deferred_interest',
        'отсроч поступления': 'deferred_interest',
        'отсроч. поступления': 'deferred_interest',
        'отсроченн поступления': 'deferred_interest',
        'отсроченн. поступления': 'deferred_interest',
        'отсроченные поступлений': 'deferred_interest',
        # Пени, штрафы, неустойки (объединенный столбец)
        'пени, штрафы, неустойки': 'penalties',
        'пени штрафы неустойки': 'penalties',
        'пени': 'penalties',
        # Старые варианты для обратной совместимости (пеня за ОД)
        'пеня за од': 'penalty_principal_old',
        'сумма пеня за од': 'penalty_principal_old',
        'сумма пени за од': 'penalty_principal_old',
        'неустойка': 'penalty_principal_old',
        'штраф': 'penalty_principal_old',
        # Старые варианты для обратной совместимости (пеня за вознаграждение)
        'пеня за вознаграждение': 'penalty_reward_old',
        'сумма пеня за вознаграждение': 'penalty_reward_old',
        'сумма пени за вознаграждение': 'penalty_reward_old',
        # Административные сборы (включая гос.пошлину)
        'административные сборы': 'admin_fees',
        'адм. сборы': 'admin_fees',
        'адм сборы': 'admin_fees',
        # Старые варианты гос.пошлины (теперь часть административных сборов)
        'гос.пошлина': 'admin_fees',
        'гос. пошлина': 'admin_fees',
        'госпошлина': 'admin_fees',
        'сумма госпошлины': 'admin_fees',
        # Общая сумма (если есть готовое значение в Excel)
        'сумма займа': 'total',
        'общая сумма': 'total',
        'итого': 'total'
    }

    col_indices = {}
    for header_name, col_idx in headers.items():
        header_lower = header_name.lower().strip()
        # Сначала ищем точное совпадение
        if header_lower in column_map:
            field_name = column_map[header_lower]
            # Пропускаем столбцы, помеченные как ignore
            if field_name != 'ignore' and field_name not in col_indices:
                col_indices[field_name] = col_idx
        else:
            # Частичное совпадение
            for key, field_name in column_map.items():
                if key in header_lower or header_lower in key:
                    # Пропускаем столбцы, помеченные как ignore
                    if field_name != 'ignore' and field_name not in col_indices:
                        col_indices[field_name] = col_idx
                    break

    clients = []
    for row in range(2, ws.max_row + 1):
        if not ws.cell(row=row, column=1).value:
            continue

        client = {
            'id': row - 1,
            'contract_number': '',
            'contract_date': '',
            'client_name': '',
            'iin': '',  # ИИН клиента
            'principal': 0,
            'reward': 0,
            'deferred_interest': 0,
            'penalties': 0,  # Пени, штрафы, неустойки (объединенные)
            'admin_fees': 0,  # Административные сборы (включая гос.пошлину)
            'total': 0,
            # Временные поля для обратной совместимости
            'penalty_principal_old': 0,
            'penalty_reward_old': 0
        }

        for field_name, col_idx in col_indices.items():
            value = ws.cell(row=row, column=col_idx).value
            if value is not None:
                if field_name == 'iin':
                    # Особая обработка ИИН для сохранения ведущих нулей
                    if isinstance(value, (int, float)):
                        # Если число, преобразуем в строку и дополняем нулями до 12 цифр
                        str_value = str(int(value)).zfill(12)
                    else:
                        # Если строка, просто убираем пробелы
                        str_value = str(value).strip()
                    client[field_name] = str_value if str_value else ''
                elif field_name in ['contract_number', 'contract_date', 'client_name']:
                    # Преобразуем в строку и убираем пробелы
                    str_value = str(value).strip()
                    client[field_name] = str_value if str_value else ''
                else:
                    try:
                        val = float(value) if value else 0
                        if field_name == 'admin_fees':
                            # Суммируем все значения из столбцов гос.пошлины и админ.сборов
                            client['admin_fees'] += val
                        else:
                            client[field_name] = val
                    except:
                        if field_name not in ['penalty_principal_old', 'penalty_reward_old']:
                            client[field_name] = 0

        # Если penalties не был задан напрямую, суммируем из старых столбцов
        if client['penalties'] == 0:
            client['penalties'] = client['penalty_principal_old'] + client['penalty_reward_old']

        # Удаляем временные поля
        del client['penalty_principal_old']
        del client['penalty_reward_old']

        # Если total не был задан из Excel, рассчитываем его
        if client['total'] == 0:
            client['total'] = (
                client['principal'] + client['reward'] + client['deferred_interest'] +
                client['penalties'] + client['admin_fees']
            )

        # Валидация ИИН
        iin_validation = validate_iin(client['iin'])
        client['iin_valid'] = iin_validation['valid']
        client['iin_error'] = iin_validation['error']

        # Пропускаем пустые строки (нет номера договора, ФИО и всех сумм)
        if not client['contract_number'] and not client['client_name'] and client['total'] == 0:
            continue

        clients.append(client)

    wb.close()
    return clients


def create_excel_certificate(client: dict, report_date: str, manager: str, output_path: Path):
    """Создание справки в формате Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Справка"

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 75
    ws.column_dimensions['C'].width = 25

    row = 1

    # Шапка с логотипом
    if LOGO_PATH.exists():
        try:
            img = XLImage(str(LOGO_PATH))
            img.width = 400
            img.height = 60
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 50
            row = 2
        except Exception as e:
            # Если не удалось вставить картинку - текстовая шапка
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = COMPANY['name']
            ws[f'A{row}'].font = Font(name='Arial', size=18, bold=True)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            ws.row_dimensions[row].height = 30
            row += 1
    else:
        # Текстовая шапка если нет логотипа
        ws.merge_cells(f'A{row}:C{row}')

        # Создаем цветной текст: "Swiss" красным, "Capital" черным
        red_font = InlineFont(rFont='Arial', sz=18, b=True, color='FF0000')
        black_font = InlineFont(rFont='Arial', sz=18, b=True, color='000000')

        rich_text = CellRichText(
            TextBlock(red_font, 'Swiss'),
            TextBlock(black_font, 'Capital')
        )

        ws[f'A{row}'].value = rich_text
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 30
        row += 1

        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = COMPANY['address']
        ws[f'A{row}'].font = Font(size=9, color='666666')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1

        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = f"телефон: {COMPANY['phone']}"
        ws[f'A{row}'].font = Font(size=9, color='666666')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1

    row += 5  # Разрыв 5 строк после шапки

    # Заголовок
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Расчет ссудной задолженности"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 5  # Разрыв 5 строк после заголовка

    # Основной текст
    contract_date = format_date_russian(client['contract_date'])
    total_text = format_number_with_text(int(client['total']))

    iin_text = f", ИИН {client['iin']}" if client.get('iin') else ""
    main_text = (
        f"По договору займа №{client['contract_number']} от {contract_date}, "
        f"Заемщик: {client['client_name']}{iin_text}, по состоянию на {report_date} "
        f"ссудная задолженность составляет {total_text} тенге, из них:"
    )

    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = main_text
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top', indent=2)
    ws.row_dimensions[row].height = 60
    row += 2

    # Детализация
    details = [
        ('Основной долг', client['principal'], True),
        ('Вознаграждение', client['reward'], True),
    ]

    if client['deferred_interest'] > 0:
        details.append(('Сумма отсроченных процентов', client['deferred_interest'], False))
    if client['penalties'] > 0:
        details.append(('Пени, штрафы, неустойки', client['penalties'], True))
    if client['admin_fees'] > 0:
        details.append(('Прочие поступления (административные сборы, гос.пошлина)', client['admin_fees'], False))

    for label, amount, show_text in details:
        ws[f'A{row}'] = "➤"
        ws[f'A{row}'].alignment = Alignment(horizontal='center', indent=2)
        ws.merge_cells(f'B{row}:C{row}')

        if show_text and amount > 0:
            text = f"{label} - {format_number_with_text(int(amount))} тенге;"
        else:
            text = f"{label} – {int(amount):,} тенге;".replace(',', ' ')

        ws[f'B{row}'] = text
        ws[f'B{row}'].alignment = Alignment(indent=2)
        row += 1

    row += 10  # Разрыв 10 строк перед подписью

    # Подпись (жирным шрифтом)
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Операционный менеджер"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = manager
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].alignment = Alignment(horizontal='right')

    wb.save(output_path)
    wb.close()


def create_pdf_certificate(client: dict, report_date: str, manager: str, output_path: Path):
    """Создание справки в формате PDF"""
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )

    # Стили с кириллическим шрифтом
    title_style = ParagraphStyle(
        'Title',
        fontName=PDF_FONT_BOLD,
        fontSize=16,  # Уменьшено с 18
        alignment=1,
        spaceAfter=4  # Уменьшено с 6
    )

    subtitle_style = ParagraphStyle(
        'Subtitle',
        fontName=PDF_FONT,
        fontSize=8,  # Уменьшено с 9
        textColor=colors.grey,
        alignment=1,
        spaceAfter=2  # Уменьшено с 3
    )

    heading_style = ParagraphStyle(
        'Heading',
        fontName=PDF_FONT_BOLD,
        fontSize=11,  # Уменьшено с 12
        alignment=1,
        spaceBefore=10,  # Уменьшено с 20
        spaceAfter=10  # Уменьшено с 20
    )

    body_style = ParagraphStyle(
        'Body',
        fontName=PDF_FONT,
        fontSize=10,
        leading=14,
        spaceAfter=10,
        leftIndent=10*mm  # Отступ 1 см
    )

    bullet_style = ParagraphStyle(
        'Bullet',
        fontName=PDF_FONT,
        fontSize=10,
        leftIndent=10*mm,  # Отступ 1 см
        spaceAfter=5
    )

    story = []

    # Шапка с цветным названием компании
    company_name_html = '<font color="red">Swiss</font>Capital'
    story.append(Paragraph(company_name_html, title_style))
    story.append(Paragraph(COMPANY['address'], subtitle_style))
    story.append(Paragraph(f"телефон: {COMPANY['phone']}", subtitle_style))
    story.append(Spacer(1, 50))  # Разрыв 5 строк после шапки

    # Заголовок
    story.append(Paragraph("Расчет ссудной задолженности", heading_style))
    story.append(Spacer(1, 50))  # Разрыв 5 строк после заголовка

    # Основной текст
    contract_date = format_date_russian(client['contract_date'])
    total_text = format_number_with_text(int(client['total']))

    iin_text = f", ИИН {client['iin']}" if client.get('iin') else ""
    main_text = (
        f"По договору займа №{client['contract_number']} от {contract_date}, "
        f"Заемщик: {client['client_name']}{iin_text}, по состоянию на {report_date} "
        f"ссудная задолженность составляет {total_text} тенге, из них:"
    )

    story.append(Paragraph(main_text, body_style))
    story.append(Spacer(1, 10))

    # Детализация
    details = [
        ('Основной долг', client['principal'], True),
        ('Вознаграждение', client['reward'], True),
    ]

    if client['deferred_interest'] > 0:
        details.append(('Сумма отсроченных процентов', client['deferred_interest'], False))
    if client['penalties'] > 0:
        details.append(('Пени, штрафы, неустойки', client['penalties'], True))
    if client['admin_fees'] > 0:
        details.append(('Прочие поступления (административные сборы, гос.пошлина)', client['admin_fees'], False))

    for label, amount, show_text in details:
        if show_text and amount > 0:
            text = f"• {label} - {format_number_with_text(int(amount))} тенге;"
        else:
            text = f"• {label} – {int(amount):,} тенге;".replace(',', ' ')
        story.append(Paragraph(text, bullet_style))

    story.append(Spacer(1, 100))  # Разрыв 10 строк перед подписью

    # Подпись (жирным шрифтом)
    signature_data = [
        ['Операционный менеджер', manager]
    ]
    signature_table = Table(signature_data, colWidths=[120*mm, 50*mm])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), PDF_FONT_BOLD),  # Жирный шрифт
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    story.append(signature_table)

    doc.build(story)


def generate_all_certificates(clients: List[dict], report_date: str, manager: str,
                               output_dir: Path, formats: List[str]) -> dict:
    """Генерация всех справок"""
    excel_dir = output_dir / "excel"
    pdf_dir = output_dir / "pdf"

    if 'excel' in formats:
        excel_dir.mkdir(exist_ok=True)
    if 'pdf' in formats:
        pdf_dir.mkdir(exist_ok=True)

    generated = {'excel': [], 'pdf': []}

    for client in clients:
        safe_name = "".join(c for c in client['client_name'] if c.isalnum() or c in ' _-').strip()[:50]
        filename = f"{client['id']:04d}_{safe_name}"

        if 'excel' in formats:
            excel_path = excel_dir / f"{filename}.xlsx"
            create_excel_certificate(client, report_date, manager, excel_path)
            generated['excel'].append(excel_path)

        if 'pdf' in formats:
            pdf_path = pdf_dir / f"{filename}.pdf"
            create_pdf_certificate(client, report_date, manager, pdf_path)
            generated['pdf'].append(pdf_path)

    return generated


def create_zip_archive(files: List[Path], output_path: Path):
    """Создание ZIP архива"""
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for file in files:
            zf.write(file, file.name)


# ============================================================================
# МАРШРУТЫ
# ============================================================================

@app.get("/", response_class=HTMLResponse)
async def home(request: Request, username: str = Depends(verify_credentials)):
    """Главная страница"""
    return templates.TemplateResponse("index.html", {
        "request": request,
        "username": username,
        "history": generation_history[-10:][::-1]
    })


@app.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    username: str = Depends(verify_credentials)
):
    """Загрузка Excel файла"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Только Excel файлы (.xlsx, .xls)")

    # Сохраняем файл
    session_id = str(uuid.uuid4())
    session_dir = UPLOAD_DIR / session_id
    session_dir.mkdir(exist_ok=True)

    file_path = session_dir / file.filename
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    # Читаем данные
    try:
        clients = read_excel_data(file_path)
        column_info = get_column_mapping_info(file_path)
    except Exception as e:
        shutil.rmtree(session_dir)
        raise HTTPException(400, f"Ошибка чтения файла: {str(e)}")

    return {
        "session_id": session_id,
        "filename": file.filename,
        "clients_count": len(clients),
        "clients": clients,
        "column_mapping": column_info
    }


@app.post("/update-client/{session_id}")
async def update_client(
    session_id: str,
    client_data: dict,
    username: str = Depends(verify_credentials)
):
    """Обновление данных клиента"""
    # В реальном приложении здесь нужно обновить данные в сессии/БД
    return {"status": "ok", "message": "Данные обновлены"}


@app.get("/preview/{session_id}/{client_id}")
async def preview_certificate(
    session_id: str,
    client_id: int,
    report_date: str,
    manager: str,
    username: str = Depends(verify_credentials)
):
    """Предпросмотр справки (HTML)"""
    session_dir = UPLOAD_DIR / session_id
    if not session_dir.exists():
        raise HTTPException(404, "Сессия не найдена")

    # Читаем данные
    files = list(session_dir.glob("*.xlsx")) + list(session_dir.glob("*.xls"))
    if not files:
        raise HTTPException(404, "Файл не найден")

    clients = read_excel_data(files[0])
    client = next((c for c in clients if c['id'] == client_id), None)

    if not client:
        raise HTTPException(404, "Клиент не найден")

    # Генерируем HTML предпросмотр
    contract_date = format_date_russian(client['contract_date'])
    total_text = format_number_with_text(int(client['total']))

    details = []
    details.append(f"Основной долг - {format_number_with_text(int(client['principal']))} тенге")
    details.append(f"Вознаграждение – {format_number_with_text(int(client['reward']))} тенге")

    if client['deferred_interest'] > 0:
        details.append(f"Сумма отсроченных процентов – {int(client['deferred_interest']):,} тенге".replace(',', ' '))
    if client['penalties'] > 0:
        details.append(f"Пени, штрафы, неустойки – {format_number_with_text(int(client['penalties']))} тенге")
    if client['admin_fees'] > 0:
        details.append(f"Прочие поступления (административные сборы, гос.пошлина) – {int(client['admin_fees']):,} тенге".replace(',', ' '))

    return {
        "company": COMPANY,
        "contract_number": client['contract_number'],
        "contract_date": contract_date,
        "client_name": client['client_name'],
        "iin": client.get('iin', ''),  # Добавляем ИИН
        "report_date": report_date,
        "total": int(client['total']),
        "total_text": total_text,
        "details": details,
        "manager": manager
    }


@app.post("/generate/{session_id}")
async def generate_certificates(
    session_id: str,
    report_date: str = Form(...),
    manager: str = Form(...),
    format_type: str = Form(...),
    username: str = Depends(verify_credentials)
):
    """Генерация справок"""
    session_dir = UPLOAD_DIR / session_id
    if not session_dir.exists():
        raise HTTPException(404, "Сессия не найдена")

    files = list(session_dir.glob("*.xlsx")) + list(session_dir.glob("*.xls"))
    if not files:
        raise HTTPException(404, "Файл не найден")

    clients = read_excel_data(files[0])

    # Создаём директорию для результатов
    output_id = str(uuid.uuid4())
    output_dir = GENERATED_DIR / output_id
    output_dir.mkdir(exist_ok=True)

    # Определяем форматы
    formats = []
    if format_type in ['excel', 'both']:
        formats.append('excel')
    if format_type in ['pdf', 'both']:
        formats.append('pdf')

    # Генерируем справки
    generated = generate_all_certificates(clients, report_date, manager, output_dir, formats)

    # Создаём архивы
    archives = {}
    if generated['excel']:
        excel_zip = output_dir / "certificates_excel.zip"
        create_zip_archive(generated['excel'], excel_zip)
        archives['excel'] = f"/download/{output_id}/excel"

    if generated['pdf']:
        pdf_zip = output_dir / "certificates_pdf.zip"
        create_zip_archive(generated['pdf'], pdf_zip)
        archives['pdf'] = f"/download/{output_id}/pdf"

    # Сохраняем в историю
    generation_history.append({
        'id': output_id,
        'date': datetime.now().strftime('%d.%m.%Y %H:%M'),
        'report_date': report_date,
        'manager': manager,
        'clients_count': len(clients),
        'formats': formats,
        'archives': archives
    })

    return {
        "status": "ok",
        "output_id": output_id,
        "clients_count": len(clients),
        "archives": archives
    }


@app.get("/download/{output_id}/{format_type}")
async def download_archive(
    output_id: str,
    format_type: str,
    username: str = Depends(verify_credentials)
):
    """Скачивание архива"""
    output_dir = GENERATED_DIR / output_id

    if format_type == 'excel':
        file_path = output_dir / "certificates_excel.zip"
        filename = "certificates_excel.zip"
    else:
        file_path = output_dir / "certificates_pdf.zip"
        filename = "certificates_pdf.zip"

    if not file_path.exists():
        raise HTTPException(404, "Файл не найден")

    return FileResponse(
        file_path,
        filename=filename,
        media_type='application/zip'
    )


@app.get("/history")
async def get_history(username: str = Depends(verify_credentials)):
    """Получение истории генераций"""
    return generation_history[-20:][::-1]


@app.get("/download-template")
async def download_template(username: str = Depends(verify_credentials)):
    """Скачивание шаблона Excel"""
    template_path = PROJECT_DIR / "data" / "clients_data.xlsx"
    if not template_path.exists():
        raise HTTPException(404, "Шаблон не найден")

    return FileResponse(
        template_path,
        filename="template_clients.xlsx",
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.get("/debug-mapping/{session_id}")
async def debug_mapping(session_id: str, username: str = Depends(verify_credentials)):
    """Отладка: показать маппинг столбцов Excel"""
    session_dir = UPLOAD_DIR / session_id
    if not session_dir.exists():
        raise HTTPException(404, "Сессия не найдена")

    # Находим первый Excel файл в директории
    excel_files = list(session_dir.glob("*.xlsx")) + list(session_dir.glob("*.xls"))
    if not excel_files:
        raise HTTPException(404, "Файл не найден в сессии")

    file_path = excel_files[0]

    # Читаем Excel
    wb = load_workbook(file_path)
    ws = wb.active

    # Собираем заголовки
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip()] = col

    # Маппинг (копируем логику из read_excel_data)
    column_map = {
        # Номер договора
        'номер договора': 'contract_number',
        'номера договора': 'contract_number',
        '№ договора': 'contract_number',
        '№ номера договора': 'contract_number',
        'договор №': 'contract_number',
        'договор': 'contract_number',
        # Игнорируем столбец с порядковым номером
        '№': 'ignore',
        'п/п': 'ignore',
        '№ п/п': 'ignore',
        # Дата договора
        'дата договора': 'contract_date',
        'даты договора': 'contract_date',
        'дата': 'contract_date',
        # ФИО
        'фио': 'client_name',
        'фио клиента': 'client_name',
        'клиент': 'client_name',
        'заемщик': 'client_name',
        # ИИН
        'иин': 'iin',
        'иин клиента': 'iin',
        'iin': 'iin',
        # Основной долг
        'основной долг': 'principal',
        'сумма од': 'principal',
        'сумма основного долга': 'principal',
        'од': 'principal',
        # Вознаграждение (проценты)
        'вознаграждение': 'reward',
        'сумма вознаграждения': 'reward',
        'сумма процентов': 'reward',
        'проценты': 'reward',
        # Отсроченные проценты / поступления
        'отсроченные проценты': 'deferred_interest',
        'сумма отсроченных процентов': 'deferred_interest',
        'отсроч проценты': 'deferred_interest',
        'отсроч. проценты': 'deferred_interest',
        'отсроченн проценты': 'deferred_interest',
        'отсроченн. проценты': 'deferred_interest',
        'отсроченные поступления': 'deferred_interest',
        'сумма отсроченных поступлений': 'deferred_interest',
        'отсроч поступления': 'deferred_interest',
        'отсроч. поступления': 'deferred_interest',
        'отсроченн поступления': 'deferred_interest',
        'отсроченн. поступления': 'deferred_interest',
        'отсроченные поступлений': 'deferred_interest',
        # Пени, штрафы, неустойки (объединенный столбец)
        'пени, штрафы, неустойки': 'penalties',
        'пени штрафы неустойки': 'penalties',
        'пени': 'penalties',
        # Старые варианты для обратной совместимости (пеня за ОД)
        'пеня за од': 'penalty_principal_old',
        'сумма пеня за од': 'penalty_principal_old',
        'сумма пени за од': 'penalty_principal_old',
        'неустойка': 'penalty_principal_old',
        'штраф': 'penalty_principal_old',
        # Старые варианты для обратной совместимости (пеня за вознаграждение)
        'пеня за вознаграждение': 'penalty_reward_old',
        'сумма пеня за вознаграждение': 'penalty_reward_old',
        'сумма пени за вознаграждение': 'penalty_reward_old',
        # Административные сборы (включая гос.пошлину)
        'административные сборы': 'admin_fees',
        'адм. сборы': 'admin_fees',
        'адм сборы': 'admin_fees',
        # Старые варианты гос.пошлины (теперь часть административных сборов)
        'гос.пошлина': 'admin_fees',
        'гос. пошлина': 'admin_fees',
        'госпошлина': 'admin_fees',
        'сумма госпошлины': 'admin_fees',
        # Общая сумма (если есть готовое значение в Excel)
        'сумма займа': 'total',
        'общая сумма': 'total',
        'итого': 'total'
    }

    # Создаем маппинг (копируем логику из read_excel_data)
    col_indices = {}
    mapping_details = []

    headers_lower = {k.lower().strip(): (k, v) for k, v in headers.items()}

    for header_lower, (header_original, col_idx) in headers_lower.items():
        matched = False
        match_type = None
        matched_key = None

        # Точное совпадение
        if header_lower in column_map:
            field_name = column_map[header_lower]
            if field_name != 'ignore' and field_name not in col_indices:
                col_indices[field_name] = col_idx
                matched = True
                match_type = "exact"
                matched_key = header_lower
        else:
            # Частичное совпадение
            for key, field_name in column_map.items():
                if key in header_lower or header_lower in key:
                    if field_name != 'ignore' and field_name not in col_indices:
                        col_indices[field_name] = col_idx
                        matched = True
                        match_type = "partial"
                        matched_key = key
                    break

        # Получаем значение из второй строки (первая строка данных)
        sample_value = ws.cell(row=2, column=col_idx).value if ws.max_row >= 2 else None

        mapping_details.append({
            "column_index": col_idx,
            "header_original": header_original,
            "header_lower": header_lower,
            "matched": matched,
            "match_type": match_type,
            "matched_key": matched_key,
            "field_name": next((k for k, v in col_indices.items() if v == col_idx), None),
            "sample_value": str(sample_value) if sample_value is not None else None
        })

    wb.close()

    return {
        "session_id": session_id,
        "total_columns": len(headers),
        "mapped_fields": len(col_indices),
        "headers": list(headers.keys()),
        "field_mapping": {field: headers.get(next((k for k, v in headers.items() if v == idx), None))
                         for field, idx in col_indices.items()},
        "details": mapping_details
    }


# ============================================================================
# ЗАПУСК
# ============================================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
