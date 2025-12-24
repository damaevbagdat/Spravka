# -*- coding: utf-8 -*-
"""
Создание шаблона Excel с актуальными заголовками
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Создаем новую книгу
wb = Workbook()
ws = wb.active
ws.title = "Данные клиентов"

# Заголовки столбцов (актуальные названия)
headers = [
    "№ Номера договора",
    "Даты договора",
    "ФИО",
    "Сумма ОД",
    "Сумма процентов",
    "Сумма отсроченных процентов",
    "Сумма пеня за ОД",
    "Сумма пени за вознаграждение",
    "Сумма госпошлины",
    "Сумма займа"
]

# Записываем заголовки
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Устанавливаем ширину столбцов
ws.column_dimensions['A'].width = 18  # № Номера договора
ws.column_dimensions['B'].width = 15  # Даты договора
ws.column_dimensions['C'].width = 30  # ФИО
ws.column_dimensions['D'].width = 15  # Сумма ОД
ws.column_dimensions['E'].width = 15  # Сумма процентов
ws.column_dimensions['F'].width = 20  # Сумма отсроченных процентов
ws.column_dimensions['G'].width = 18  # Сумма пеня за ОД
ws.column_dimensions['H'].width = 25  # Сумма пени за вознаграждение
ws.column_dimensions['I'].width = 18  # Сумма госпошлины
ws.column_dimensions['J'].width = 15  # Сумма займа

# Высота строки заголовков
ws.row_dimensions[1].height = 30

# Пример данных (2 строки)
example_data = [
    ["12345", "01.01.2023", "Иванов Иван Иванович", 1000000, 150000, 0, 50000, 25000, 10000, 1235000],
    ["12346", "15.02.2023", "Петрова Мария Петровна", 500000, 75000, 10000, 25000, 12500, 5000, 627500]
]

for row_idx, data in enumerate(example_data, start=2):
    for col_idx, value in enumerate(data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Сохраняем
wb.save("data/clients_data_new.xlsx")
print("Шаблон создан: data/clients_data_new.xlsx")
