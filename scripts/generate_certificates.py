# -*- coding: utf-8 -*-
"""
Генератор справок о ссудной задолженности
SwissCapital - Республика Казахстан

Использование:
    python generate_certificates.py --date "22.12.2025" --format excel
    python generate_certificates.py --date "22.12.2025" --format pdf
    python generate_certificates.py --date "22.12.2025" --format both
"""

import os
import sys
import argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Ошибка: Установите openpyxl: pip install openpyxl")
    sys.exit(1)

from num2text import number_to_text, format_number_with_text

# Пути
SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"
OUTPUT_DIR = PROJECT_DIR / "output"
ASSETS_DIR = PROJECT_DIR / "assets"
LOGO_PATH = PROJECT_DIR / "assetslogo.png"  # Обновленный путь к логотипу

# Фиксированные данные
COMPANY_INFO = {
    'name': 'swisscapital',
    'address': 'Республика Казахстан, г.Алматы, 050026',
    'address2': 'ул. Нурмакова д.93А. тел +7(727)3550771.'
}

DEFAULT_MANAGER = "Койбасова Е.Б."


def format_date_russian(date_str):
    """Форматировать дату в русском формате"""
    months = {
        1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля',
        5: 'мая', 6: 'июня', 7: 'июля', 8: 'августа',
        9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
    }

    try:
        if isinstance(date_str, datetime):
            dt = date_str
        else:
            # Попробуем разные форматы
            for fmt in ['%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']:
                try:
                    dt = datetime.strptime(str(date_str), fmt)
                    break
                except ValueError:
                    continue
            else:
                return str(date_str)

        return f"{dt.day} {months[dt.month]} {dt.year} г."
    except:
        return str(date_str)


def read_clients_data(file_path):
    """Прочитать данные клиентов из Excel файла"""
    wb = load_workbook(file_path)
    ws = wb.active

    # Найти заголовки
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            headers[cell_value.strip().lower()] = col

    # Маппинг колонок
    column_map = {
        # Игнорируем столбец с порядковым номером
        '№': 'ignore',
        'п/п': 'ignore',
        '№ п/п': 'ignore',
        # Номер договора
        'номер договора': 'contract_number',
        'номера договора': 'contract_number',
        '№ договора': 'contract_number',
        '№ номера договора': 'contract_number',
        # Дата договора
        'дата договора': 'contract_date',
        'даты договора': 'contract_date',
        # ФИО
        'фио': 'client_name',
        'фио клиента': 'client_name',
        # Основной долг
        'основной долг': 'principal',
        'сумма од': 'principal',
        'од': 'principal',
        # Вознаграждение (проценты)
        'вознаграждение': 'reward',
        'сумма вознаграждения': 'reward',
        'сумма процентов': 'reward',
        'проценты': 'reward',
        # Отсроченные проценты
        'отсроченные проценты': 'deferred_interest',
        'сумма отсроченных процентов': 'deferred_interest',
        # Пеня за ОД
        'пеня за од': 'penalty_principal',
        'сумма пеня за од': 'penalty_principal',
        'сумма пени за од': 'penalty_principal',
        # Пеня за вознаграждение
        'пеня за вознаграждение': 'penalty_reward',
        'сумма пеня за вознаграждение': 'penalty_reward',
        'сумма пени за вознаграждение': 'penalty_reward',
        # Гос.пошлина
        'гос.пошлина': 'state_fee',
        'госпошлина': 'state_fee',
        'сумма госпошлины': 'state_fee',
        # Административные сборы
        'административные сборы': 'admin_fees',
        'адм. сборы': 'admin_fees',
        # Общая сумма
        'сумма займа': 'total',
        'общая сумма': 'total',
        'итого': 'total'
    }

    # Найти соответствия
    col_indices = {}
    for header_name, col_idx in headers.items():
        for key, field_name in column_map.items():
            if key in header_name.lower():
                # Пропускаем столбцы, помеченные как ignore
                if field_name != 'ignore' and field_name not in col_indices:
                    col_indices[field_name] = col_idx
                break

    # Читаем данные
    clients = []
    for row in range(2, ws.max_row + 1):
        # Проверяем что строка не пустая
        if not ws.cell(row=row, column=1).value:
            continue

        client = {
            'contract_number': '',
            'contract_date': '',
            'client_name': '',
            'principal': 0,
            'reward': 0,
            'deferred_interest': 0,
            'penalty_principal': 0,
            'penalty_reward': 0,
            'state_fee': 0,
            'admin_fees': 0,
            'total': 0
        }

        for field_name, col_idx in col_indices.items():
            value = ws.cell(row=row, column=col_idx).value
            if value is not None:
                if field_name in ['contract_number', 'contract_date', 'client_name']:
                    # Преобразуем в строку и убираем пробелы
                    str_value = str(value).strip()
                    client[field_name] = str_value if str_value else ''
                else:
                    try:
                        client[field_name] = float(value) if value else 0
                    except (ValueError, TypeError):
                        client[field_name] = 0

        # Вычисляем итого, если не было задано из Excel
        if client['total'] == 0:
            client['total'] = (
                client['principal'] +
                client['reward'] +
                client['deferred_interest'] +
                client['penalty_principal'] +
                client['penalty_reward'] +
                client['state_fee'] +
                client['admin_fees']
            )

        # Пропускаем пустые строки (нет номера договора, ФИО и всех сумм)
        if not client['contract_number'] and not client['client_name'] and client['total'] == 0:
            continue

        clients.append(client)

    wb.close()
    return clients


def create_certificate_excel(client, report_date, manager_name, output_path):
    """Создать справку в формате Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Справка"

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 20

    # Настройка высоты строк
    for i in range(1, 25):
        ws.row_dimensions[i].height = 18

    # Шрифты
    font_header = Font(name='Arial', size=11, bold=True)
    font_normal = Font(name='Arial', size=10)
    font_small = Font(name='Arial', size=9, color='666666')
    font_title = Font(name='Arial', size=12, bold=True)
    font_swiss = Font(name='Arial', size=16, bold=False, color='FF0000')
    font_capital = Font(name='Arial', size=16, bold=True, color='000000')

    # Выравнивание
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')

    row = 1

    # === ШАПКА ===
    # Логотип (текстовый вариант)
    ws.merge_cells(f'A{row}:C{row}')
    ws.row_dimensions[row].height = 30
    cell = ws[f'A{row}']
    cell.value = "swiss"
    cell.font = font_swiss
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Добавим "capital" в ту же ячейку через rich text эмуляцию
    # Так как openpyxl не поддерживает rich text напрямую, сделаем текстом
    ws[f'A{row}'] = ""
    row += 1

    # Текстовый логотип
    ws.merge_cells(f'A{row-1}:C{row-1}')
    ws[f'A{row-1}'].value = "swisscapital"
    ws[f'A{row-1}'].font = Font(name='Arial', size=18, bold=True)
    ws[f'A{row-1}'].alignment = align_center

    # Адрес компании
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = COMPANY_INFO['address']
    ws[f'A{row}'].font = font_small
    ws[f'A{row}'].alignment = align_center
    row += 1

    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = COMPANY_INFO['address2']
    ws[f'A{row}'].font = font_small
    ws[f'A{row}'].alignment = align_center
    row += 2

    # === ЗАГОЛОВОК ===
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = "Расчет ссудной задолженности"
    ws[f'A{row}'].font = font_title
    ws[f'A{row}'].alignment = align_center
    row += 2

    # === ОСНОВНОЙ ТЕКСТ ===
    contract_date_formatted = format_date_russian(client['contract_date'])
    report_date_formatted = format_date_russian(report_date)

    total_formatted = format_number_with_text(int(client['total']))

    main_text = (
        f"По договору займа №{client['contract_number']} от {contract_date_formatted}, "
        f"Заемщик: {client['client_name']}, по состоянию на {report_date_formatted} "
        f"ссудная задолженность составляет {total_formatted} тенге, из них:"
    )

    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = main_text
    ws[f'A{row}'].font = font_normal
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[row].height = 60
    row += 2

    # === ДЕТАЛИЗАЦИЯ ===
    # Обязательные поля
    details = [
        ('Основной долг', client['principal'], True),
        ('Вознаграждение', client['reward'], True),
    ]

    # Условные поля (показываем только если есть значение)
    if client['deferred_interest'] > 0:
        details.append(('Сумма отсроченных процентов', client['deferred_interest'], False))
    if client['penalty_principal'] > 0:
        details.append(('Неустойка (штраф за несвоевременное погашение)', client['penalty_principal'], False))
    if client['penalty_reward'] > 0:
        details.append(('Пеня за вознаграждение', client['penalty_reward'], False))
    if client['state_fee'] > 0:
        details.append(('Гос.пошлина', client['state_fee'], False))
    if client['admin_fees'] > 0:
        details.append(('Административные сборы', client['admin_fees'], False))

    for label, amount, show_text in details:
        ws[f'A{row}'].value = "➤"
        ws[f'A{row}'].alignment = align_center

        if show_text and amount > 0:
            amount_str = format_number_with_text(int(amount))
            text = f"{label} - {amount_str} тенге;"
        else:
            formatted_amount = f"{int(amount):,}".replace(',', ' ')
            text = f"{label} – {formatted_amount} тенге;"

        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'].value = text
        ws[f'B{row}'].font = font_normal
        ws[f'B{row}'].alignment = align_left
        row += 1

    row += 3

    # === ПОДПИСЬ ===
    ws[f'A{row}'].value = "Операционный менеджер"
    ws[f'A{row}'].font = font_normal
    ws.merge_cells(f'A{row}:B{row}')

    ws[f'C{row}'].value = manager_name
    ws[f'C{row}'].font = font_normal
    ws[f'C{row}'].alignment = align_right

    # Настройка печати
    ws.print_title_rows = None
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    # Сохранение
    wb.save(output_path)
    wb.close()


def generate_certificates(data_file, report_date, output_format='excel', manager_name=DEFAULT_MANAGER):
    """
    Генерировать справки для всех клиентов

    Args:
        data_file: путь к файлу с данными клиентов
        report_date: дата отчёта (строка в формате DD.MM.YYYY)
        output_format: формат вывода ('excel', 'pdf', 'both')
        manager_name: имя операционного менеджера
    """
    # Создаём папку для вывода
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Читаем данные
    print(f"Чтение данных из: {data_file}")
    clients = read_clients_data(data_file)
    print(f"Найдено клиентов: {len(clients)}")

    if not clients:
        print("Ошибка: Нет данных для обработки")
        return

    # Создаём подпапку для текущей партии
    date_folder = report_date.replace('.', '-')
    batch_dir = OUTPUT_DIR / f"batch_{date_folder}"
    batch_dir.mkdir(parents=True, exist_ok=True)

    # Генерируем справки
    for i, client in enumerate(clients, 1):
        # Формируем имя файла
        safe_name = "".join(c for c in client['client_name'] if c.isalnum() or c in ' _-').strip()
        safe_name = safe_name[:50]  # Ограничиваем длину
        filename = f"{i:04d}_{safe_name}"

        # Безопасный вывод для консоли Windows
        try:
            display_name = client['client_name'].encode('cp1251', errors='replace').decode('cp1251')
            print(f"[{i}/{len(clients)}] Создание справки: {display_name}")
        except:
            print(f"[{i}/{len(clients)}] Создание справки: Client #{i}")

        if output_format in ['excel', 'both']:
            excel_path = batch_dir / f"{filename}.xlsx"
            create_certificate_excel(client, report_date, manager_name, excel_path)

        if output_format in ['pdf', 'both']:
            # PDF генерация требует дополнительных библиотек
            # Пока создаём Excel, который можно сохранить как PDF
            print(f"   [!] PDF: Сохраните Excel как PDF или используйте Excel+VBA версию")

    print(f"\nГотово! Справки сохранены в: {batch_dir}")
    print(f"Всего создано справок: {len(clients)}")


def main():
    parser = argparse.ArgumentParser(
        description='Генератор справок о ссудной задолженности SwissCapital'
    )
    parser.add_argument(
        '--data', '-d',
        default=str(DATA_DIR / 'clients_data.xlsx'),
        help='Путь к файлу с данными клиентов (по умолчанию: data/clients_data.xlsx)'
    )
    parser.add_argument(
        '--date', '-t',
        required=True,
        help='Дата отчёта в формате DD.MM.YYYY (например: 22.12.2025)'
    )
    parser.add_argument(
        '--format', '-f',
        choices=['excel', 'pdf', 'both'],
        default='excel',
        help='Формат вывода (по умолчанию: excel)'
    )
    parser.add_argument(
        '--manager', '-m',
        default=DEFAULT_MANAGER,
        help=f'ФИО операционного менеджера (по умолчанию: {DEFAULT_MANAGER})'
    )

    args = parser.parse_args()

    # Проверяем существование файла данных
    data_path = Path(args.data)
    if not data_path.exists():
        print(f"Ошибка: Файл не найден: {data_path}")
        print(f"Создайте файл с данными клиентов или укажите другой путь через --data")
        sys.exit(1)

    generate_certificates(
        data_file=data_path,
        report_date=args.date,
        output_format=args.format,
        manager_name=args.manager
    )


if __name__ == '__main__':
    main()
